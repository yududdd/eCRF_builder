__author__ = "Yu Du"
__Email__ = "yu.du@clinchoice.com"
__date__ = "Nov 30,2020"
#######################################################################################################################
import openpyxl
import pandas as pd
from db_reader import DB
from openpyxl import Workbook
from copy import copy


def create_sheet(book, name):
    new_sheet = book.create_sheet(name)
    wb_lib = openpyxl.load_workbook('./doc/'+name+'.xlsx')
    sheet = wb_lib.get_sheet_by_name(name)
    for row in sheet:
        for cell in row:
            new_sheet[cell.coordinate].value = cell.value
            if cell.has_style:
                new_sheet[cell.coordinate].font = copy(cell.font)
                new_sheet[cell.coordinate].border = copy(cell.border)
                new_sheet[cell.coordinate].fill = copy(cell.fill)
                new_sheet[cell.coordinate].number_format = copy(cell.number_format)
                new_sheet[cell.coordinate].protection = copy(cell.protection)
                new_sheet[cell.coordinate].alignment = copy(cell.alignment)
    for index, col in sheet.column_dimensions.items():
        new_sheet.column_dimensions[index] = copy(col)


class Compiler(object):
    def __init__(self):
        self.selected = dict()
        self.db = DB()
        self.na_dict = []
        self.out_path = './output/ecrf_out.xlsx'
        self.constants = ["Signature", "Instruction", "Version", "Dynamic", "Derivation"]

    def read_selected(self, file_path):
        """
        read an excel file and return a dictionary of form and fields
        return: dictionary format of form and list in the exsiting working file
        """
        res_dict = dict()
        excel = pd.ExcelFile(file_path)
        sheets = excel.sheet_names
        for sheet in sheets:
            if sheet not in self.constants:
                df = pd.read_excel(file_path, sheet)
                fields = list(df['FieldOID'])
                res_dict[sheet] = fields
                self.selected[sheet] = fields
        return self.selected

    def delete_form(self):
        pass

    def out_text(self):
        with open("./output/Missing_Dict.txt","w") as filehandle:
            for listitem in self.na_dict:
                filehandle.write('%s\n' % listitem)
        return None

    def output(self, sign, ins, ver, dyn, der):
        book = Workbook()
        dict_df = None

        if sign:
            create_sheet(book, self.constants[0])
        if ins:
            create_sheet(book, self.constants[1])
        if ver:
            create_sheet(book, self.constants[2])
        if dyn:
            create_sheet(book, self.constants[3])
        if der:
            create_sheet(book, self.constants[4])


        fir = book.get_sheet_by_name('Sheet')
        book.remove_sheet(fir)
        with pd.ExcelWriter(self.out_path, engine='openpyxl') as writer:
            writer.book = book
            for form, fields in self.selected.items(): # first for-loop to loop through all the selected fields within a form
                df = self.db.get_selectedFields(fields)
                df.to_excel(writer, sheet_name=form, index=False)
                start_col = 0
                unique_dict = set() # data structure to store the unique dictionaies in the selected form
                for _,dict_name in df['Dictionary Name'].iteritems(): # second for-loop to loop through all the dictionaries within selected fields
                    if dict_name is not None and dict_name not in unique_dict:
                        dict_name = dict_name.replace(' ', '') # processor to replace potential trailing spaces in the dictionary string.
                        unique_dict.add(dict_name.strip()) # Use set to store all unique dictionaries
            
                for dict_name in unique_dict: #within this big loop to iterate through all selected forms.
                    dict_df = self.db.get_dict(dict_name) # get the dictionary dataframe
                    if dict_df is not None:
                        dict_df.rename({'User Data String':dict_name}, axis=1, inplace=True) # rename the dictionary header to the dictionary name
                        # try and catch to find out the error key
                        dict_df = dict_df[dict_name] # get the user data string
                        dict_df.to_excel(writer, sheet_name=form, startrow=len(df)+3, startcol= start_col, index=False) # skip one row and add dictionarys
                        start_col = start_col + 2 # skip 1 columns to write a new dictionary, plus 1 is to start a new row and plus one is skip one
            self.na_dict = self.db.na_dict # assign unfindable dictionaries
        #         writer.save()
        # writer.close()
